import string
import tkinter.ttk
import openpyxl
import socket

import Dialog
import Thermal as Th
import Constant as Cons

from tkinter import *


# Set element(label, text field, button) as specified position and size
def make_element(x, y, h, w, element, *args, **kwargs):
    f = Frame(height=h, width=w)
    f.pack_propagate(0)  # don't shrink
    f.place(x=x, y=y)
    if element.lower() == 'label':
        label = Label(f, *args, **kwargs)
        label.pack(fill=BOTH, expand=1)
        return label
    elif element.lower() == 'entry':
        text_field = Entry(f, *args, **kwargs)
        text_field.pack(fill=BOTH, expand=1)
        return text_field
    elif element.lower() == 'button':
        btn = Button(f, *args, **kwargs)
        btn.pack(fill=BOTH, expand=1)
        return btn


# Set Command Table
def make_table(root: tkinter, column_num: int, width: int, column_title: [string], x: int, y: int, cmd_data: []):
    # Create a table, column is name of column,
    # The displaycolumn shows the order in which the table is executed.

    host = Cons.host_ip
    input_port = Cons.port
    port = int(0) if Cons.port == '' else int(input_port)
    buf_size = Cons.buf_size

    column = [i + 1 for i in range(column_num)]
    dis_column = [str(n) for n in column]
    tv = tkinter.ttk.Treeview(root, columns=column, displaycolumns=dis_column)
    # Treeview의 width, height 글자 수로 정해 진다.
    tv.configure(height=len(Cons.command_array) + 1)
    tv.place(x=x, y=y)

    # Set the Table No Tab
    tv.column('#0', width=50, anchor='center', stretch='yes')
    tv.heading('#0', text='No', anchor='center')

    # Create each column(name, width, anchor)
    for i in range(column_num):
        tv.column(dis_column[i], width=width, anchor='center')
        tv.heading(dis_column[i], text=column_title[i], anchor='center')

    # Insert the command data in Table
    for i in range(len(cmd_data)):
        tv.insert('', 'end', text=i, values=cmd_data[i], iid=str(i) + '번')
        # CMD Title Left Align
        tv.column(dis_column[0], anchor='w')
        # tv.bind('<<TreeviewSelect>>', selectItem)
        tv.bind('<<TreeviewSelect>>', lambda event, root_view=root: check_network_Info(event, root_view))


def selectItem(event, root_view):
    hex_array = []
    tree = event.widget
    # selection = [tree.item(item)["text"] for item in tree.selection()]
    selection = [tree.item(item)['values'][1] for item in tree.selection()]
    print(selection)
    for i in range(0, len(selection[0]), 2):
        hex_str = '0x' + selection[0][i:i + 2]
        hex_int = int(hex_str, 16)
        hex_array.append(hex_int)

    print(hex_array)

    Th.send_data(hex_array, root_view)


def check_network_Info(event, root_view):
    host = Cons.host_ip
    input_port = Cons.port
    port = int(0) if Cons.port == '' else int(input_port)

    if host != "" and port != 0:
        selectItem(event, root_view)
    else:
        print("Invalid command")
        # network_notification()
        dialog_txt = 'Please enter a network info.'
        dialog = Dialog.DialogBox(root_view, dialog_txt)


# Get the Command from csv file
def get_data_from_csv(file_path) -> [(str, str)]:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sh = wb.worksheets[0]
    sh = wb['IR RAW CMD']
    sh = wb.active

    max_column = sh.max_column
    max_row = sh.max_row

    command_data = []
    for i, row in enumerate(sh.iter_rows(max_col=max_column - 2, max_row=max_row - 2), start=3):
        cmd_title = sh.cell(row=i, column=2).value
        cmd_data = sh.cell(row=i, column=15).value
        cmd_pair = (cmd_title, cmd_data)
        command_data.append(cmd_pair)

    return command_data


# TODO: Generate exe format

# TODO: CheckBox
# TODO: Interval between command
# TODO: Command send priority
# TODO: Check whether sent command was applied

# TODO: PTZ UI (Rudder, Focus, Zoom)
# TODO: Absolute move
# TODO: Move to point of click (Centering)

# TODO: RTSP












